#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    :2019年12月29日 
# @Author  : ghost-guest
# @Site    : 
# @File    : 合并对个相同表头但纵向单元格合并的execl文件.py
# @Software: PyCharm
# 说明： 合并单元格，要求如标题需求


from os import listdir
from os.path import exists
import os
import openpyxl
# 结果文件名，如果已经存在，先删除
result = './dir/result.xlsx'
if exists(result):
    os.remove(result)

# 创建空白结果文件，并添加表头
wb_result = openpyxl.Workbook()
ws_result = wb_result.worksheets[0]
ws_result.append(['学院', '姓名', '成绩'])
# 遍历当前文件夹中所有xlsx文件
# 把除去表头之外的内容追加到结果文件中
print(os.getcwd())
print(listdir('/Users/mac/Desktop/python/dir'))
fns = ['/Users/mac/Desktop/python/dir/'+fn for fn in listdir('/Users/mac/Desktop/python/dir') if fn.endswith('.xlsx')]
print(fns)
print("="*20)
for fn in fns:
    wb = openpyxl.load_workbook(fn)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows):
        # 跳过表头
        if index == 0:
            continue
        ws_result.append(list(map(lambda cell: cell.value, row)))
        print("------:", ws_result)
# 结果文件中所有行，前面加一个空串，方便索引
print('ws_result:', ws_result)
wb_result.save(result)
rows = [''] + list(ws_result.rows)
print('修改后的列表：', rows)
index1 = 2
rowCount = len(rows)
print('-'*20)
print(rowCount)
# 处理结果文件，合并第一列中合适的单元格
# while index1 < rowCount:
for i in range(rowCount):
    value = rows[index1][0].value
    # 如果当前单元格没有内容，或者与前面的内容相同，就合并
    print('========', value)
    print(rows[index1+1:])
    for index2, row2 in enumerate(rows[index1+1:], start=index1+1):
        print(index2, row2)
        print(row2[0].value)
        if row2[0].value == None:
            ws_result.merge_cells('A' + str(index1) + ':A' + str(index2))
            #break
        elif row2[0].value == value:
            # 已经到文件尾，合并单元格
            ws_result.merge_cells('A' + str(index1) + ':A' + str(index2))
            print("1223---23124=====:")
            #break
        # 未到文件尾，合并单元格
            #ws_result.merge_cells('A' + str(index1) + ':A' + str(index2 - 1))
        index1 = index2
        print(index1)

# 保存结果文件
wb_result.save(result)